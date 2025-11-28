"""
utils/excel_handler.py

Improved Excel handler with robust text measurement (works across Pillow versions).

Outputs:
 - per-sheet annotated PNGs
 - annotated_combined.pdf
 - annotated_<original>.xlsx (cells with red border + wrapped text & adjusted row heights)

Usage:
    python utils/excel_handler.py path/to/file.xlsx "query" --outdir output

Dependencies:
    pip install openpyxl pillow
"""
import os
import sys
from pathlib import Path
import traceback
import textwrap
from math import ceil

from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment

# ensure project root importable to reuse images_to_pdf from pdf_handler
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from pdf_handler import images_to_pdf
except Exception as e:
    raise ImportError("Could not import images_to_pdf from pdf_handler.py. Ensure pdf_handler.py is in project root.") from e

# -----------------------------
# Constants & helpers
# -----------------------------
CHAR_PX = 7
CELL_PADDING_PX = 8
DEFAULT_CELL_HEIGHT = 30
MIN_COL_PX = 60
MAX_COL_PX = 800
MAX_COL_CHARS = 100

def choose_font(font_size=14):
    try:
        return ImageFont.truetype("arial.ttf", font_size)
    except Exception:
        return ImageFont.load_default()

def pixel_width_from_chars(n_chars, char_px=CHAR_PX, pad=CELL_PADDING_PX):
    w = int(n_chars * char_px + pad * 2)
    return max(MIN_COL_PX, min(w, MAX_COL_PX))

def wrap_text_for_width(text, max_chars_per_line):
    wrapper = textwrap.TextWrapper(width=max_chars_per_line, break_long_words=True, replace_whitespace=False)
    lines = []
    for paragraph in str(text).splitlines():
        if paragraph.strip() == "":
            lines.append("")
        else:
            wrapped = wrapper.wrap(paragraph)
            if not wrapped:
                lines.append("")
            else:
                lines.extend(wrapped)
    return lines

# Robust text measurement that works across Pillow versions
def measure_text(draw: ImageDraw.ImageDraw, font: ImageFont.ImageFont, text: str):
    """
    Return (width_px, height_px) for given text using best available method.
    """
    text = "" if text is None else str(text)
    # 1) draw.textbbox
    try:
        bbox = draw.textbbox((0, 0), text, font=font)
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        return int(w), int(h)
    except Exception:
        pass

    # 2) draw.textlength + font.getmetrics (for height estimate)
    try:
        if hasattr(draw, "textlength"):
            w = int(draw.textlength(text, font=font))
        else:
            raise AttributeError
        # try font.getmetrics for ascent/descent
        if hasattr(font, "getmetrics"):
            ascent, descent = font.getmetrics()
            h = int(ascent + descent)
        else:
            # fallback approximate
            h = int(font.size if hasattr(font, "size") else 14)
        return w, h
    except Exception:
        pass

    # 3) font.getsize (older versions)
    try:
        size = font.getsize(text)
        if isinstance(size, tuple) and len(size) >= 2:
            return int(size[0]), int(size[1])
    except Exception:
        pass

    # 4) approximate fallback
    avg_char_w = getattr(font, "size", 7) * 0.6
    w = int(len(text) * avg_char_w)
    h = getattr(font, "size", 14)
    return w, h

# -----------------------------
# Excel autosize helpers
# -----------------------------
def autosize_sheet_excel(ws, min_col_width_chars=10, max_col_width_chars=60):
    max_len_per_col = {}
    for row in ws.iter_rows(values_only=True):
        for c_idx, val in enumerate(row, start=1):
            if val is None:
                continue
            s = str(val)
            l = max(len(line) for line in s.splitlines()) if s else 0
            max_len_per_col[c_idx] = max(max_len_per_col.get(c_idx, 0), l)

    for c_idx, max_len in max_len_per_col.items():
        col_letter = get_column_letter(c_idx)
        width_chars = max(min_col_width_chars, min(max_len + 2, max_col_width_chars))
        ws.column_dimensions[col_letter].width = width_chars

def set_row_heights_for_wrapped_cells(ws, font_size=14):
    """
    Estimate row heights based on wrapped text and set ws.row_dimensions heights (points).
    Uses robust measurement via PIL.
    """
    font = choose_font(font_size)
    dummy_img = Image.new("RGB", (10, 10))
    draw = ImageDraw.Draw(dummy_img)

    for r in range(1, ws.max_row + 1):
        max_lines = 1
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            col_letter = get_column_letter(c)
            col_width_chars = ws.column_dimensions[col_letter].width or 10
            max_chars = max(1, int(col_width_chars))
            lines = wrap_text_for_width(str(val), max_chars)
            max_lines = max(max_lines, max(1, len(lines)))
        # measure line height
        _, line_h = measure_text(draw, font, "Ay")
        approx_px_row = max_lines * (line_h + 4)
        approx_pts = approx_px_row * 0.75
        ws.row_dimensions[r].height = approx_pts

# -----------------------------
# Render sheet to image (improved formatting)
# -----------------------------
def render_sheet_to_image(ws, query, font_size=14):
    font = choose_font(font_size)
    draw_dummy = ImageDraw.Draw(Image.new("RGB", (10, 10)))

    col_count = ws.max_column or 1
    col_px = []
    max_chars_per_col = []
    for c in range(1, col_count + 1):
        col_letter = get_column_letter(c)
        width_chars = ws.column_dimensions[col_letter].width or 10
        width_chars = min(width_chars, MAX_COL_CHARS)
        px = pixel_width_from_chars(width_chars)
        col_px.append(px)
        max_chars_per_col.append(max(1, int(width_chars)))

    row_count = ws.max_row or 1
    row_px = []
    for r in range(1, row_count + 1):
        max_lines = 1
        for c in range(1, col_count + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            lines = wrap_text_for_width(str(val), max_chars_per_col[c-1])
            max_lines = max(max_lines, len(lines))
        _, line_h = measure_text(draw_dummy, font, "Ay")
        row_height_px = max(DEFAULT_CELL_HEIGHT, max_lines * (line_h + 4) + 8)
        row_px.append(row_height_px)

    img_w = sum(col_px)
    img_h = sum(row_px)
    img = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(img)

    y = 0
    cell_boxes = {}
    for r in range(1, row_count + 1):
        x = 0
        for c in range(1, col_count + 1):
            w_px = col_px[c-1]
            h_px = row_px[r-1]
            x0, y0 = x, y
            x1, y1 = x + w_px, y + h_px
            cell_boxes[(r, c)] = (x0, y0, x1, y1)
            # border
            draw.rectangle([x0, y0, x1, y1], outline="black")
            val = ws.cell(row=r, column=c).value
            text = "" if val is None else str(val)
            if text != "":
                lines = wrap_text_for_width(text, max_chars_per_col[c-1])
                _, line_h = measure_text(draw, font, "Ay")
                total_h = len(lines) * (line_h + 4)
                ty = y0 + max(4, (h_px - total_h) / 2)
                for line in lines:
                    tw, _ = measure_text(draw, font, line)
                    tx = x0 + max(4, (w_px - tw) / 2)
                    draw.text((tx, ty), line, fill="black", font=font)
                    ty += line_h + 4
            x = x1
        y += row_px[r-1]

    # find matches
    matches = find_matches_in_sheet(ws, query)
    return img, cell_boxes, matches

# -----------------------------
# Find matches in sheet
# -----------------------------
def find_matches_in_sheet(ws, query):
    q = query.lower().strip()
    matches = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            try:
                if q in str(val).lower():
                    matches.append((r, c))
            except Exception:
                continue
    return matches

# -----------------------------
# Draw matches on image
# -----------------------------
def draw_matches_on_image(img, cell_boxes, match_cells, out_path, stroke="red", stroke_width=4):
    draw = ImageDraw.Draw(img)
    for (r, c) in match_cells:
        if (r, c) not in cell_boxes:
            continue
        x0, y0, x1, y1 = cell_boxes[(r, c)]
        for w in range(stroke_width):
            draw.rectangle([x0 - w, y0 - w, x1 + w, y1 + w], outline=stroke)
    img.save(out_path)
    return True

# -----------------------------
# Top-level processor
# -----------------------------
def process_excel_file(xlsx_path, query, out_dir="output",
                       stroke_color="red", outline_width=4, font_size=14):
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"[excel_handler] Input Excel : {xlsx_path}")
    print(f"[excel_handler] Query       : '{query}'")
    print(f"[excel_handler] Output dir  : {out_dir.resolve()}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    annotated_wb = openpyxl.load_workbook(xlsx_path)

    generated_images = []

    # needed openpyxl style objects
    red_side = Side(border_style="thin", color="FF0000")
    red_border = Border(left=red_side, right=red_side, top=red_side, bottom=red_side)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"[excel_handler] Processing sheet: {sheet_name}")

        annotated_ws = annotated_wb[sheet_name]
        autosize_sheet_excel(annotated_ws)
        set_row_heights_for_wrapped_cells(annotated_ws, font_size=font_size)

        match_cells = find_matches_in_sheet(ws, query)
        print(f"[excel_handler]  - matches found: {len(match_cells)}")

        # mark matches in annotated workbook
        for (r, c) in match_cells:
            cell = annotated_ws.cell(row=r, column=c)
            cell.border = red_border
            if cell.alignment is None:
                cell.alignment = Alignment(wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal=cell.alignment.horizontal or None,
                                           vertical=cell.alignment.vertical or None,
                                           wrap_text=True)

        try:
            img, cell_boxes, matches = render_sheet_to_image(annotated_ws, query, font_size=font_size)
        except Exception as e:
            print(f"[excel_handler] ERROR rendering sheet '{sheet_name}': {e}")
            traceback.print_exc()
            continue

        out_img_path = out_dir / f"{sheet_name}_annot.png"
        draw_matches_on_image(img, cell_boxes, match_cells, out_img_path, stroke=stroke_color, stroke_width=outline_width)
        print(f"[excel_handler]  - saved annotated image: {out_img_path}")
        generated_images.append(str(out_img_path))

    annotated_name = out_dir / f"annotated_{xlsx_path.name}"
    try:
        annotated_wb.save(annotated_name)
        print(f"[excel_handler] Annotated Excel saved: {annotated_name}")
    except Exception as e:
        print(f"[excel_handler] ERROR saving annotated Excel: {e}")
        traceback.print_exc()

    combined_pdf = out_dir / "annotated_combined.pdf"
    if generated_images:
        try:
            images_to_pdf(generated_images, combined_pdf)
            print(f"[excel_handler] Combined PDF saved at: {combined_pdf}")
        except Exception as e:
            print(f"[excel_handler] ERROR creating combined PDF: {e}")
            traceback.print_exc()
    else:
        print("[excel_handler] No annotated images generated, skipping combined PDF.")

    return generated_images, str(combined_pdf), str(annotated_name)

# -----------------------------
# CLI
# -----------------------------
if __name__ == "__main__":
    import argparse
    from openpyxl.styles import Side, Border, Alignment

    parser = argparse.ArgumentParser(description="Excel annotation handler")
    parser.add_argument("xlsx", help="Path to Excel (.xlsx) file")
    parser.add_argument("query", help="Query text to search in cells")
    parser.add_argument("--outdir", default="output", help="Output folder")
    parser.add_argument("--outline", type=int, default=4, help="Box outline width (px)")
    parser.add_argument("--color", default="red", help="Box color")
    parser.add_argument("--font-size", type=int, default=14, help="Font size for image rendering")

    args = parser.parse_args()

    try:
        images, combined, annotated_xlsx = process_excel_file(
            args.xlsx,
            args.query,
            out_dir=args.outdir,
            stroke_color=args.color,
            outline_width=args.outline,
            font_size=args.font_size
        )

        print("\nOutput summary:")
        if images:
            print("Images:")
            for p in images:
                print(" -", p)
        if combined:
            print("Combined PDF:", combined)
        if annotated_xlsx:
            print("Annotated Excel:", annotated_xlsx)

    except Exception as e:
        print("[excel_handler] ERROR while processing Excel file:")
        traceback.print_exc()
        sys.exit(1)
